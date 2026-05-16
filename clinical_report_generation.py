"""Generate per-case reports for the keloid treatment classification pipeline."""

import argparse
import html
import os
from pathlib import Path


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}


def log(message):
    print(f"[Reports] {message}", flush=True)


def safe_id(name):
    return Path(name).stem.replace(" ", "_")


def build_image_index(root):
    root = Path(root)
    if not root.exists():
        return {}

    index = {}
    for file in root.rglob("*"):
        if file.is_file() and file.suffix.lower() in IMAGE_EXTENSIONS:
            index.setdefault(file.stem, file)
            index.setdefault(file.name, file)
    return index


def get_prediction_rows(prediction_xlsx):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Report generation requires openpyxl. "
            "Install project dependencies with: pip install -r requirements.txt"
        ) from exc

    prediction_xlsx = Path(prediction_xlsx)
    if not prediction_xlsx.exists():
        raise FileNotFoundError(f"Missing prediction table: {prediction_xlsx}")

    workbook = load_workbook(prediction_xlsx, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value) if value is not None else "" for value in rows[0]]
    required_columns = {"name", "Conservative", "Aggressive"}
    missing_columns = required_columns - set(headers)
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"Prediction table is missing columns: {missing}")

    records = []
    for values in rows[1:]:
        record = dict(zip(headers, values))
        if record.get("name"):
            records.append(record)
    return records


def get_case_names(prediction_rows, crop_dir):
    if prediction_rows:
        return [str(row["name"]) for row in prediction_rows]

    crop_dir = Path(crop_dir)
    if not crop_dir.exists():
        return []

    return sorted(
        file.name
        for file in crop_dir.iterdir()
        if file.is_file() and file.suffix.lower() in IMAGE_EXTENSIONS
    )


def find_image(image_index, name):
    path = image_index.get(name) or image_index.get(Path(name).stem)
    return Path(path) if path else None


def relpath(path, output_file):
    return Path(os.path.relpath(Path(path).resolve(), output_file.parent.resolve())).as_posix()


def image_panel(title, path, output_file):
    title = html.escape(title)
    if path is None or not Path(path).exists():
        return f"""
        <section class="panel missing">
          <h2>{title}</h2>
          <div class="placeholder">Not available</div>
        </section>
        """

    src = html.escape(relpath(path, output_file))
    alt = html.escape(Path(path).name)
    return f"""
    <section class="panel">
      <h2>{title}</h2>
      <img src="{src}" alt="{alt}">
    </section>
    """


def format_probability(value):
    return f"{float(value):.3f}"


def prediction_block(row):
    conservative = float(row.get("Conservative", 0.0))
    aggressive = float(row.get("Aggressive", 0.0))
    label = "Aggressive" if aggressive >= conservative else "Conservative"
    confidence = max(conservative, aggressive)
    conservative_pct = max(0.0, min(100.0, conservative * 100.0))
    aggressive_pct = max(0.0, min(100.0, aggressive * 100.0))

    return f"""
    <section class="prediction">
      <div>
        <span class="label">Recommended treatment class</span>
        <strong>{html.escape(label)}</strong>
      </div>
      <div>
        <span class="label">Confidence</span>
        <strong>{confidence:.3f}</strong>
      </div>
      <div class="probability">
        <div class="probability-row">
          <span>Conservative</span>
          <span>{format_probability(conservative)}</span>
        </div>
        <div class="bar"><span style="width: {conservative_pct:.1f}%"></span></div>
        <div class="probability-row">
          <span>Aggressive</span>
          <span>{format_probability(aggressive)}</span>
        </div>
        <div class="bar aggressive"><span style="width: {aggressive_pct:.1f}%"></span></div>
      </div>
    </section>
    """


def render_case_report(case_name, row, paths, output_file):
    output_file.parent.mkdir(parents=True, exist_ok=True)

    escaped_name = html.escape(case_name)
    panels = "\n".join(
        [
            image_panel("Original clinical image", paths["raw"], output_file),
            image_panel("Cropped lesion region", paths["crop"], output_file),
            image_panel("Mask overlay", paths["fuse"], output_file),
            image_panel("Predicted blood perfusion", paths["perfusion"], output_file),
        ]
    )

    output_file.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Keloid Blood Perfusion Report - {escaped_name}</title>
  <style>
    body {{
      margin: 0;
      background: #f6f7f9;
      color: #20242a;
      font-family: Arial, Helvetica, sans-serif;
    }}
    main {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 32px 24px 48px;
    }}
    header {{
      border-bottom: 1px solid #d9dde3;
      margin-bottom: 24px;
      padding-bottom: 18px;
    }}
    h1 {{
      font-size: 28px;
      margin: 0 0 8px;
      letter-spacing: 0;
    }}
    .case-id {{
      color: #5b6470;
      font-size: 14px;
    }}
    .prediction {{
      display: grid;
      grid-template-columns: minmax(180px, 1fr) minmax(160px, 1fr) minmax(260px, 2fr);
      gap: 20px;
      align-items: start;
      background: #ffffff;
      border: 1px solid #dde2e8;
      border-radius: 8px;
      padding: 18px;
      margin-bottom: 24px;
    }}
    .label {{
      display: block;
      color: #65707d;
      font-size: 13px;
      margin-bottom: 6px;
    }}
    strong {{
      font-size: 22px;
    }}
    .probability-row {{
      display: flex;
      justify-content: space-between;
      color: #303741;
      font-size: 14px;
      margin-bottom: 6px;
    }}
    .bar {{
      height: 8px;
      background: #edf0f4;
      border-radius: 4px;
      margin-bottom: 12px;
      overflow: hidden;
    }}
    .bar span {{
      display: block;
      height: 100%;
      background: #336f8a;
    }}
    .bar.aggressive span {{
      background: #a44e3f;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 18px;
    }}
    .panel {{
      background: #ffffff;
      border: 1px solid #dde2e8;
      border-radius: 8px;
      padding: 14px;
    }}
    .panel h2 {{
      font-size: 15px;
      margin: 0 0 12px;
      color: #303741;
    }}
    .panel img {{
      width: 100%;
      aspect-ratio: 1 / 1;
      object-fit: contain;
      background: #eef1f5;
      border-radius: 6px;
      display: block;
    }}
    .placeholder {{
      display: grid;
      place-items: center;
      aspect-ratio: 1 / 1;
      background: #eef1f5;
      color: #6d7784;
      border-radius: 6px;
    }}
    footer {{
      color: #6d7784;
      font-size: 12px;
      margin-top: 28px;
    }}
    @media (max-width: 760px) {{
      main {{
        padding: 24px 16px 36px;
      }}
      .prediction {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <h1>Keloid Blood Perfusion Report</h1>
      <div class="case-id">Case: {escaped_name}</div>
    </header>
    {prediction_block(row)}
    <section class="grid">
      {panels}
    </section>
    <footer>Generated automatically from the keloid treatment classification pipeline.</footer>
  </main>
</body>
</html>
""",
        encoding="utf-8",
    )


def render_index(case_reports, output_file):
    rows = []
    for item in case_reports:
        case_name = html.escape(item["case_name"])
        report_path = html.escape(relpath(item["report"], output_file))
        conservative = format_probability(item["row"]["Conservative"])
        aggressive = format_probability(item["row"]["Aggressive"])
        label = "Aggressive" if float(item["row"]["Aggressive"]) >= float(item["row"]["Conservative"]) else "Conservative"
        rows.append(
            f"""
            <tr>
              <td><a href="{report_path}">{case_name}</a></td>
              <td>{html.escape(label)}</td>
              <td>{conservative}</td>
              <td>{aggressive}</td>
            </tr>
            """
        )

    output_file.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Keloid Blood Perfusion Reports</title>
  <style>
    body {{
      margin: 0;
      background: #f6f7f9;
      color: #20242a;
      font-family: Arial, Helvetica, sans-serif;
    }}
    main {{
      max-width: 980px;
      margin: 0 auto;
      padding: 32px 24px 48px;
    }}
    h1 {{
      font-size: 28px;
      margin: 0 0 20px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      background: #ffffff;
      border: 1px solid #dde2e8;
      border-radius: 8px;
      overflow: hidden;
    }}
    th, td {{
      border-bottom: 1px solid #e5e9ee;
      padding: 12px 14px;
      text-align: left;
      font-size: 14px;
    }}
    th {{
      background: #eef1f5;
      color: #303741;
    }}
    tr:last-child td {{
      border-bottom: 0;
    }}
    a {{
      color: #275f7a;
      text-decoration: none;
      font-weight: 600;
    }}
  </style>
</head>
<body>
  <main>
    <h1>Keloid Blood Perfusion Reports</h1>
    <table>
      <thead>
        <tr>
          <th>Case</th>
          <th>Recommended treatment class</th>
          <th>Conservative</th>
          <th>Aggressive</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows)}
      </tbody>
    </table>
  </main>
</body>
</html>
""",
        encoding="utf-8",
    )


def parse_args():
    parser = argparse.ArgumentParser(description="Generate per-case keloid treatment reports.")
    parser.add_argument("--raw-path", default="./data/raw_data")
    parser.add_argument("--crop-path", default="./data/crop_result")
    parser.add_argument("--fuse-path", default="./data/fuse_result")
    parser.add_argument("--perfusion-result-path", default="./data/perfusion_result")
    parser.add_argument("--prediction-xlsx", default="output_prediction.xlsx")
    parser.add_argument("--output-dir", default="./reports")
    return parser.parse_args()


def main():
    args = parse_args()
    output_dir = Path(args.output_dir)
    case_dir = output_dir / "cases"
    output_dir.mkdir(parents=True, exist_ok=True)

    prediction_rows = get_prediction_rows(args.prediction_xlsx)
    rows_by_name = {str(row["name"]): row for row in prediction_rows}

    indexes = {
        "raw": build_image_index(args.raw_path),
        "crop": build_image_index(args.crop_path),
        "fuse": build_image_index(args.fuse_path),
        "perfusion": build_image_index(args.perfusion_result_path),
    }

    case_names = get_case_names(prediction_rows, args.crop_path)
    if not case_names:
        raise FileNotFoundError("No cases were found for report generation.")

    case_reports = []
    for case_name in case_names:
        row = rows_by_name.get(case_name)
        if row is None:
            log(f"Skipped {case_name}: prediction row not found")
            continue

        paths = {key: find_image(index, case_name) for key, index in indexes.items()}
        output_file = case_dir / f"{safe_id(case_name)}.html"
        render_case_report(case_name, row, paths, output_file)
        case_reports.append({"case_name": case_name, "row": row, "report": output_file})

    if not case_reports:
        raise RuntimeError("No reports were generated.")

    index_file = output_dir / "index.html"
    render_index(case_reports, index_file)
    log(f"Case reports: {case_dir}")
    log(f"Report index: {index_file}")
    log(f"Completed; reports generated: {len(case_reports)}")


if __name__ == "__main__":
    main()
